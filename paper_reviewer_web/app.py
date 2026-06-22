"""FastAPI app for the Paper Reviewer web dashboard.

Run from the project root:
    python -m paper_reviewer_web.app

Then open http://localhost:8765 (or whatever port the launcher prints).
"""

from __future__ import annotations

import asyncio
import json
import os
import shutil
import sys
import threading
import time
import traceback
import webbrowser
from pathlib import Path
from typing import Optional

from fastapi import (
    FastAPI, File, Form, HTTPException, Request, UploadFile, BackgroundTasks,
)
from fastapi.responses import (
    HTMLResponse, JSONResponse, StreamingResponse, RedirectResponse, FileResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jinja2 import Environment, FileSystemLoader, select_autoescape
import uvicorn

# Make sibling 'paper_reviewer' package importable when running by file path
_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE.parent))

from paper_reviewer import corpus as corpus_mod, excel_io, extractor, plagiarism, reviewer, rules as rules_mod, templates as review_templates
from paper_reviewer.providers import (
    ProviderManager, GroqProvider, OllamaProvider, OpenRouterProvider, CONFIG_PATH,
)
from paper_reviewer_web.tasks import STATE, PaperItem, upload_dir, new_paper_id


app = FastAPI(title="Paper Reviewer Dashboard")

# When frozen by PyInstaller, templates/static are unpacked under sys._MEIPASS.
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    _ASSET_BASE = Path(sys._MEIPASS) / "paper_reviewer_web"
else:
    _ASSET_BASE = _HERE
TEMPLATES_DIR = _ASSET_BASE / "templates"
STATIC_DIR = _ASSET_BASE / "static"

# cache_size=0 sidesteps a Jinja2 LRUCache / Python 3.14 dict-key TypeError.
_jinja_env = Environment(
    loader=FileSystemLoader(str(TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
    cache_size=0,
)
templates = Jinja2Templates(env=_jinja_env)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ---------------- Pages ----------------


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse(
        request, "index.html",
        {"default_excel": str(Path.cwd() / "reviews_output.xlsx")},
    )


@app.get("/check", response_class=HTMLResponse)
def check_page(request: Request):
    return templates.TemplateResponse(request, "check.html", {})


@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request):
    mgr = ProviderManager.load()
    return templates.TemplateResponse(
        request, "settings.html",
        {
            "providers": mgr.to_dict()["providers"],
            "auto_fallback": mgr.auto_fallback,
            "config_path": str(CONFIG_PATH),
        },
    )


# ---------------- File uploads ----------------


@app.post("/api/upload")
async def upload_papers(files: list[UploadFile] = File(...)):
    out = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in (".pdf", ".docx"):
            continue
        pid = new_paper_id()
        dest = upload_dir() / f"{pid}{ext}"
        with dest.open("wb") as fp:
            shutil.copyfileobj(f.file, fp)
        guess_id = extractor.extract_paper_id(f.filename or "") or ""
        item = PaperItem(
            id=pid,
            filename=f.filename or "",
            server_path=str(dest),
            paper_id=guess_id,
        )
        STATE.items.append(item)
        out.append({
            "id": item.id,
            "filename": item.filename,
            "paper_id": item.paper_id,
        })
    return {"items": out}


@app.post("/api/cmt-upload")
async def cmt_upload(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(400, "CMT file must be .xlsx or .xls")
    dest = upload_dir() / f"cmt{ext}"
    with dest.open("wb") as fp:
        shutil.copyfileobj(file.file, fp)
    try:
        md = excel_io.load_cmt_metadata(str(dest))
    except Exception as e:
        raise HTTPException(400, f"Could not parse CMT file: {e}")
    STATE.cmt_path = str(dest)
    STATE.cmt_count = len(md)
    return {"count": len(md), "ids": sorted(md.keys())[:50]}


@app.post("/api/items/{item_id}")
async def update_item(item_id: str, request: Request):
    body = await request.json()
    for item in STATE.items:
        if item.id == item_id:
            if "paper_id" in body:
                item.paper_id = str(body["paper_id"]).strip()
            if "force_reject" in body:
                v = (body.get("force_reject") or "").strip()
                if v not in ("", "ai", "plag", "novelty"):
                    raise HTTPException(400, "bad force_reject")
                item.force_reject = v
            return {"ok": True}
    raise HTTPException(404, "item not found")


@app.delete("/api/items/{item_id}")
async def remove_item(item_id: str):
    before = len(STATE.items)
    STATE.items = [i for i in STATE.items if i.id != item_id]
    return {"removed": before - len(STATE.items)}


@app.post("/api/items/clear")
async def clear_items():
    STATE.items = []
    return {"ok": True}


@app.get("/api/items")
async def list_items():
    return {
        "items": [
            {
                "id": i.id, "filename": i.filename, "paper_id": i.paper_id,
                "force_reject": i.force_reject, "status": i.status,
                "opinion": i.opinion, "review": i.review, "provider": i.provider,
                "error": i.error,
            }
            for i in STATE.items
        ],
        "cmt_path": STATE.cmt_path,
        "cmt_count": STATE.cmt_count,
    }


# ---------------- Job control ----------------


@app.post("/api/job/start")
async def start_job(request: Request):
    body = await request.json()
    STATE.excel_path = body.get("excel_path") or str(Path.cwd() / "reviews_output.xlsx")
    STATE.preview_enabled = bool(body.get("preview_enabled", True))
    if STATE.running:
        raise HTTPException(409, "A job is already running")
    if not STATE.items:
        raise HTTPException(400, "No papers uploaded")
    mgr = ProviderManager.load()
    if not any(p.enabled for p in mgr.providers.values()):
        raise HTTPException(400, "No provider is enabled. Open Settings and enable at least one.")

    STATE.running = True
    STATE.cancelled = False
    t = threading.Thread(target=_run_job, args=(mgr,), daemon=True)
    t.start()
    return {"started": True, "count": len(STATE.items)}


@app.post("/api/job/cancel")
async def cancel_job():
    STATE.cancelled = True
    # release any pending preview so worker can notice cancellation
    STATE.preview_response = {"action": "cancel_all"}
    STATE.preview_event.set()
    return {"ok": True}


@app.post("/api/job/preview-response")
async def preview_response(request: Request):
    body = await request.json()
    action = body.get("action", "accept")
    review = body.get("review", "")
    opinion = body.get("opinion", "")
    STATE.preview_response = {"action": action, "review": review, "opinion": opinion}
    STATE.preview_event.set()
    return {"ok": True}


@app.get("/api/job/stream")
async def job_stream():
    async def gen():
        # Always send a hello so the client knows it's connected.
        yield _sse({"type": "hello", "data": {"running": STATE.running}})
        while True:
            try:
                ev = STATE.events.get(timeout=0.5)
                yield _sse(ev)
                if ev["type"] == "job_done":
                    break
            except Exception:
                # keep-alive
                yield ": ping\n\n"
                if not STATE.running and STATE.events.empty():
                    # nothing more coming; close
                    break
            await asyncio.sleep(0)
    return StreamingResponse(gen(), media_type="text/event-stream")


def _sse(ev: dict) -> str:
    return f"event: {ev['type']}\ndata: {json.dumps(ev['data'])}\n\n"


def _run_job(mgr: ProviderManager):
    try:
        wb, sheet = excel_io.ensure_workbook(STATE.excel_path)
    except Exception as e:
        STATE.push("job_error", {"message": f"Could not open Excel: {e}"})
        STATE.push("job_done", {})
        STATE.running = False
        return

    cmt_md = {}
    if STATE.cmt_path:
        try:
            cmt_md = excel_io.load_cmt_metadata(STATE.cmt_path)
        except Exception as e:
            STATE.push("log", {"message": f"CMT metadata load failed: {e}"})

    total = len(STATE.items)
    for idx, item in enumerate(STATE.items, 1):
        if STATE.cancelled:
            item.status = "cancelled"
            STATE.push("item_update", _item_dict(item))
            continue
        try:
            item.status = "reading"
            STATE.push("item_update", _item_dict(item))
            text = extractor.extract_text(item.server_path)
            title_hint = extractor.guess_title(text)
            page_count = extractor.count_pages(item.server_path)
            sections = extractor.detect_sections(text, path=item.server_path)
            STATE.push("log", {
                "message": f"[{item.paper_id or item.filename}] {page_count} pages, "
                           f"present: {', '.join(s for s, ok in sections.items() if ok) or '(none)'}"
            })

            # Plagiarism check (skip if already checked OR user already forced a reject)
            plag_cfg = rules_mod.load_rules().get("global", {}).get("plagiarism", {})
            if (plag_cfg.get("auto_check_before_review", True)
                    and item.ai_score < 0 and item.dup_score < 0
                    and not item.force_reject):
                item.status = "plagiarism"
                STATE.push("item_update", _item_dict(item))
                try:
                    plag_result = _run_plagiarism_for_item(item)
                    STATE.push("log", {
                        "message": f"[{item.paper_id or item.filename}] plagiarism: "
                                   f"AI={plag_result.ai_score}% dup={plag_result.dup_score}% "
                                   f"{'FLAGGED→'+plag_result.flagged if plag_result.flagged else 'OK'}"
                    })
                    if plag_result.flagged and plag_cfg.get("auto_flag_reject_if_over", True):
                        item.force_reject = plag_result.flagged
                except Exception as e:
                    STATE.push("log", {"message": f"[plagiarism check failed] {e}"})

            meta = cmt_md.get(item.paper_id, {}) if item.paper_id else {}
            title_for_row = meta.get("paper_title") or title_hint

            item.status = "reviewing"
            STATE.push("item_update", _item_dict(item))

            reason = item.force_reject or None
            result = reviewer.review_paper(
                mgr, text, title_hint=title_for_row, force_reject_reason=reason,
                page_count=page_count, sections=sections,
            )
            item.review = result.review
            item.opinion = result.opinion
            item.provider = result.provider

            # Preview step
            if STATE.preview_enabled:
                item.status = "preview"
                STATE.push("item_update", _item_dict(item))
                STATE.pending_preview_id = item.id
                STATE.preview_event.clear()
                STATE.push("preview", {
                    "id": item.id,
                    "filename": item.filename,
                    "paper_id": item.paper_id,
                    "title": title_for_row,
                    "review": item.review,
                    "opinion": item.opinion,
                    "provider": item.provider,
                })
                STATE.preview_event.wait()
                resp = STATE.preview_response
                STATE.preview_response = {}
                STATE.pending_preview_id = ""
                action = resp.get("action", "accept")
                if action == "cancel_all":
                    STATE.cancelled = True
                    item.status = "cancelled"
                    STATE.push("item_update", _item_dict(item))
                    continue
                if action == "skip":
                    item.status = "skipped"
                    STATE.push("item_update", _item_dict(item))
                    continue
                # accept (possibly edited)
                if resp.get("review"):
                    item.review = resp["review"]
                if resp.get("opinion"):
                    item.opinion = resp["opinion"]

            row = {
                "paper_id": item.paper_id,
                "paper_title": title_for_row,
                "primary_name": meta.get("primary_name", ""),
                "primary_email": meta.get("primary_email", ""),
                "authors": meta.get("authors", ""),
                "author_names": meta.get("author_names", ""),
                "author_emails": meta.get("author_emails", ""),
                "review": item.review,
                "opinion": item.opinion,
            }
            excel_io.append_row(wb, sheet, row)
            excel_io.save(wb, STATE.excel_path)
            item.status = "done"
            STATE.push("item_update", _item_dict(item))

            # Add to the persistent corpus so future uploads are checked against this one
            try:
                corpus_mod.add_paper(
                    text, filename=item.filename, paper_id=item.paper_id,
                    title=title_for_row,
                )
            except Exception as e:
                STATE.push("log", {"message": f"[corpus add failed] {e}"})

        except Exception as e:
            item.status = "error"
            item.error = str(e)
            STATE.push("item_update", _item_dict(item))
            traceback.print_exc()

        STATE.push("progress", {"current": idx, "total": total})

    STATE.push("job_done", {"excel": STATE.excel_path})
    STATE.running = False


def _item_dict(item: PaperItem) -> dict:
    return {
        "id": item.id, "filename": item.filename, "paper_id": item.paper_id,
        "force_reject": item.force_reject, "status": item.status,
        "opinion": item.opinion, "review": item.review,
        "provider": item.provider, "error": item.error,
        "ai_score": item.ai_score, "dup_score": item.dup_score,
        "web_score": item.web_score, "similarity_index": item.similarity_index,
        "plag_flagged": item.plag_flagged, "plag_evidence": item.plag_evidence,
        "last_report": item.last_report,
    }


def _run_plagiarism_for_item(item: PaperItem, *, deep: bool = False) -> plagiarism.PlagiarismResult:
    """Extract text for this paper + build a session corpus from OTHER in-flight uploads.
    The persistent corpus (corpus_mod) is queried automatically inside check_paper().
    """
    text = extractor.extract_text(item.server_path)
    session_corpus: list[tuple[str, str]] = []
    for other in STATE.items:
        if other.id == item.id:
            continue
        try:
            session_corpus.append((other.filename or other.id, extractor.extract_text(other.server_path)))
        except Exception:
            pass
    cfg = rules_mod.load_rules().get("global", {}).get("plagiarism", {})
    ai_t = float(cfg.get("ai_threshold", 50))
    dup_t = float(cfg.get("dup_threshold", 50))
    result = plagiarism.check_paper(
        text, session_corpus,
        ai_threshold=ai_t, dup_threshold=dup_t,
        deep_web_check=deep,
        gptzero_api_key=(cfg.get("gptzero_api_key") or "").strip(),
    )
    item.ai_score = result.ai_score
    item.dup_score = result.dup_score
    item.web_score = result.web_score
    item.similarity_index = result.similarity_index
    item.plag_flagged = result.flagged or ""
    item.plag_evidence = (result.ai_evidence + result.dup_evidence + result.web_evidence)
    item.last_report = result.to_dict()
    return result


# ---------------- Results & stats ----------------


@app.get("/api/results")
async def results(excel_path: Optional[str] = None):
    path = excel_path or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    p = Path(path)
    if not p.exists():
        return {"rows": [], "stats": {}, "excel_path": str(p)}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = []
        stats: dict[str, int] = {}
        excel_row = 1
        for r in ws.iter_rows(min_row=2, values_only=True):
            excel_row += 1
            if not any(r):
                continue
            row = {
                "excel_row": excel_row,  # 1-based Excel row (incl. header at row 1)
                "paper_id": r[0] if len(r) > 0 else "",
                "paper_title": r[1] if len(r) > 1 else "",
                "primary_name": r[2] if len(r) > 2 else "",
                "review": r[7] if len(r) > 7 else "",
                "opinion": r[8] if len(r) > 8 else "",
            }
            rows.append(row)
            op = (row["opinion"] or "").strip() or "(blank)"
            stats[op] = stats.get(op, 0) + 1
        wb.close()
        return {"rows": rows, "stats": stats, "excel_path": str(p)}
    except Exception as e:
        return {"rows": [], "stats": {}, "excel_path": str(p), "error": str(e)}


@app.post("/api/results/delete")
async def delete_result(request: Request):
    body = await request.json()
    path = body.get("excel_path") or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    row_index = body.get("excel_row")
    if not row_index:
        raise HTTPException(400, "excel_row is required")
    try:
        ok = excel_io.delete_row(path, int(row_index))
    except Exception as e:
        raise HTTPException(400, str(e))
    return {"deleted": ok, "excel_row": row_index}


@app.post("/api/results/clear")
async def clear_results(request: Request):
    body = await request.json()
    path = body.get("excel_path") or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    try:
        n = excel_io.clear_all_rows(path)
    except Exception as e:
        raise HTTPException(400, str(e))
    return {"deleted": n}


@app.post("/api/cache/clear")
async def clear_review_cache():
    n = reviewer.clear_cache()
    return {"cleared": n}


# ---------------- Plagiarism check ----------------


@app.post("/api/plagiarism/{item_id}")
async def run_plagiarism(item_id: str):
    item = next((i for i in STATE.items if i.id == item_id), None)
    if item is None:
        raise HTTPException(404, "item not found")
    try:
        result = _run_plagiarism_for_item(item, deep=False)
    except Exception as e:
        raise HTTPException(500, str(e))
    cfg = rules_mod.load_rules().get("global", {}).get("plagiarism", {})
    if result.flagged and cfg.get("auto_flag_reject_if_over", True):
        item.force_reject = result.flagged
    return result.to_dict() | {"force_reject": item.force_reject}


@app.post("/api/plagiarism/{item_id}/deep")
async def run_plagiarism_deep(item_id: str):
    """Same as /api/plagiarism but also queries DuckDuckGo for ~6 distinctive
    sentences. Slow (5-20s) and network-dependent."""
    item = next((i for i in STATE.items if i.id == item_id), None)
    if item is None:
        raise HTTPException(404, "item not found")
    try:
        result = _run_plagiarism_for_item(item, deep=True)
    except Exception as e:
        raise HTTPException(500, str(e))
    cfg = rules_mod.load_rules().get("global", {}).get("plagiarism", {})
    if result.flagged and cfg.get("auto_flag_reject_if_over", True):
        item.force_reject = result.flagged
    return result.to_dict() | {"force_reject": item.force_reject}


@app.post("/api/check")
async def standalone_check(files: list[UploadFile] = File(...), deep: bool = False):
    """Standalone AI + plagiarism check — does NOT touch the review queue, does NOT
    call the LLM, does NOT write Excel. Each paper is compared against the persistent
    corpus + the other files in this same upload batch. `deep=true` adds a web search.
    """
    cfg = rules_mod.load_rules().get("global", {}).get("plagiarism", {})
    ai_t = float(cfg.get("ai_threshold", 50))
    dup_t = float(cfg.get("dup_threshold", 50))
    gpt_key = (cfg.get("gptzero_api_key") or "").strip()

    # Extract all texts first (for cross-comparison within the batch)
    papers: list[tuple[str, str]] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in (".pdf", ".docx"):
            continue
        pid = new_paper_id()
        dest = upload_dir() / f"check_{pid}{ext}"
        with dest.open("wb") as fp:
            shutil.copyfileobj(f.file, fp)
        try:
            text = extractor.extract_text(str(dest))
        except Exception:
            text = ""
        papers.append((f.filename or pid, text))

    results = []
    for i, (fname, text) in enumerate(papers):
        if not text:
            results.append({"filename": fname, "error": "could not extract text"})
            continue
        session_corpus = [(papers[j][0], papers[j][1]) for j in range(len(papers)) if j != i]
        r = plagiarism.check_paper(
            text, session_corpus,
            ai_threshold=ai_t, dup_threshold=dup_t,
            deep_web_check=deep, gptzero_api_key=gpt_key,
        )
        results.append({"filename": fname, **r.to_dict()})
    return {"results": results, "deep": deep}


# ---------------- Corpus admin ----------------


@app.get("/api/corpus")
async def list_corpus():
    entries = [
        {
            "sha": e.sha, "filename": e.filename, "paper_id": e.paper_id,
            "title": e.title, "added_at": e.added_at, "word_count": e.word_count,
        }
        for e in corpus_mod.list_corpus()
    ]
    return {"entries": entries, "count": len(entries), "dir": str(corpus_mod.CORPUS_DIR)}


@app.delete("/api/corpus/{sha}")
async def delete_corpus_entry(sha: str):
    return {"removed": corpus_mod.remove_paper(sha)}


@app.post("/api/corpus/clear")
async def clear_corpus():
    n = corpus_mod.clear_corpus()
    return {"cleared": n}


# ---------------- Publisher rules (Review Settings) ----------------


@app.get("/api/opinions")
async def get_opinions():
    return {"opinions": reviewer.allowed_opinions()}


@app.get("/api/publisher-rules")
async def get_publisher_rules():
    return {
        "rules": rules_mod.load_rules(),
        "rules_path": str(rules_mod.RULES_PATH),
        "rules_hash": rules_mod.rules_hash(),
    }


@app.post("/api/publisher-rules")
async def save_publisher_rules(request: Request):
    body = await request.json()
    new_rules = body.get("rules")
    if not new_rules or "publishers" not in new_rules or "global" not in new_rules:
        raise HTTPException(400, "Body must be {rules: {publishers: [...], global: {...}}}")
    # Lightweight schema check
    for p in new_rules["publishers"]:
        if not isinstance(p.get("name"), str) or not p["name"].strip():
            raise HTTPException(400, "Each publisher needs a non-empty name")
        if not isinstance(p.get("rules"), list):
            raise HTTPException(400, f"Publisher '{p.get('name')}' has no rules list")
        try:
            int(p.get("min_matches", 1))
        except Exception:
            raise HTTPException(400, f"Publisher '{p.get('name')}' has invalid min_matches")
    rules_mod.save_rules(new_rules)
    # Invalidate the review cache since rules changed
    reviewer.clear_cache()
    return {"saved": True, "rules_hash": rules_mod.rules_hash()}


@app.post("/api/publisher-rules/restore-defaults")
async def restore_publisher_defaults():
    fresh = rules_mod.restore_defaults()
    reviewer.clear_cache()
    return {"rules": fresh, "rules_hash": rules_mod.rules_hash()}


@app.get("/api/excel-status")
async def excel_status(excel_path: Optional[str] = None):
    path = excel_path or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    p = Path(path)
    return {
        "exists": p.exists(),
        "size": p.stat().st_size if p.exists() else 0,
        "modified": p.stat().st_mtime if p.exists() else None,
        "path": str(p),
        "name": p.name,
    }


@app.get("/api/download")
async def download_excel(excel_path: Optional[str] = None):
    path = excel_path or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    p = Path(path)
    if not p.exists():
        raise HTTPException(404, f"Excel file not found at {p}. Process at least one paper first.")
    return FileResponse(
        path=str(p),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=p.name,
    )


@app.post("/api/open-folder")
async def open_folder(request: Request):
    body = await request.json()
    path = body.get("excel_path") or STATE.excel_path or str(Path.cwd() / "reviews_output.xlsx")
    p = Path(path)
    target = p if p.exists() else p.parent
    if not target.exists():
        raise HTTPException(404, "Folder not found")
    try:
        if sys.platform == "win32":
            if p.exists():
                # opens Explorer with the file selected
                os.system(f'explorer /select,"{p}"')
            else:
                os.startfile(str(target))
        elif sys.platform == "darwin":
            os.system(f'open "{target}"')
        else:
            os.system(f'xdg-open "{target}"')
    except Exception as e:
        raise HTTPException(500, f"Could not open folder: {e}")
    return {"ok": True, "opened": str(target)}


# ---------------- Provider settings ----------------


@app.get("/api/providers")
async def get_providers():
    mgr = ProviderManager.load()
    return {
        "providers": mgr.to_dict()["providers"],
        "order": mgr.order,
        "auto_fallback": mgr.auto_fallback,
        "config_path": str(CONFIG_PATH),
    }


@app.post("/api/providers")
async def save_providers(request: Request):
    body = await request.json()
    mgr = ProviderManager.load()
    for name, cfg in body.get("providers", {}).items():
        if name not in mgr.providers:
            continue
        prov = mgr.providers[name]
        prov.enabled = bool(cfg.get("enabled", False))
        if cfg.get("model"):
            prov.model = str(cfg["model"]).strip()
        if isinstance(prov, (GroqProvider, OpenRouterProvider)):
            keys = cfg.get("api_keys", [])
            if isinstance(keys, str):
                keys = [k.strip() for k in keys.splitlines() if k.strip()]
            prov.api_keys = [k for k in keys if k]
            prov._key_cycle = None
        if isinstance(prov, OllamaProvider):
            url = cfg.get("base_url")
            if url:
                prov.base_url = url.strip()
    mgr.auto_fallback = bool(body.get("auto_fallback", True))
    mgr.save()
    return {"saved": True, "config_path": str(CONFIG_PATH)}


@app.post("/api/providers/{name}/test")
async def test_provider(name: str):
    mgr = ProviderManager.load()
    p = mgr.providers.get(name)
    if not p:
        raise HTTPException(404, "unknown provider")
    ok, msg = p.test()
    return {"ok": ok, "message": msg}


# ---------------- Launcher ----------------


def main():
    port = int(os.environ.get("PAPER_REVIEWER_PORT", "8765"))
    url = f"http://localhost:{port}"
    print(f"\n  Paper Reviewer Dashboard")
    print(f"  Open in browser:  {url}\n")
    try:
        webbrowser.open(url)
    except Exception:
        pass
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning")


if __name__ == "__main__":
    main()
