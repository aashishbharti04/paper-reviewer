"""Excel I/O — read existing CMT-style sheets and append review rows.

Output Excel uses the same 9-column schema as the user's existing ICCCNet2026 file:
  Paper ID | Paper Title | Primary Contact Author Name | Primary Contact Author Email |
  Authors | Author Names | Author Emails | Review | Opinion

If author metadata isn't available (e.g. when working purely from a PDF), those
fields are left blank — the user can paste them later from the CMT export.
"""

from __future__ import annotations

import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


CMT_FIELDS = {
    "Paper ID": "paper_id",
    "Paper Title": "paper_title",
    "Primary Contact Author Name": "primary_name",
    "Primary Contact Author Email": "primary_email",
    "Authors": "authors",
    "Author Names": "author_names",
    "Author Emails": "author_emails",
}


HEADERS = [
    "Paper ID",
    "Paper Title",
    "Primary Contact Author Name",
    "Primary Contact Author Email",
    "Authors",
    "Author Names",
    "Author Emails",
    "Review",
    "Opinion",
]


def _header_style():
    return {
        "font": Font(name="Arial", bold=True),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", start_color="D9E1F2"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def _row_style():
    return {
        "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def ensure_workbook(path: str) -> tuple[Workbook, str]:
    """Open the workbook at `path`, or create a new one with proper headers.

    Returns (workbook, sheet_name).
    """
    p = Path(path)
    if p.exists() and p.suffix.lower() == ".xlsx":
        wb = load_workbook(path)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        # If sheet is empty, add headers
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            _write_headers(ws)
        return wb, sheet_name
    # Fresh workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"
    _write_headers(ws)
    return wb, ws.title


def _write_headers(ws) -> None:
    style = _header_style()
    for col, header in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = style["font"]
        c.alignment = style["alignment"]
        c.fill = style["fill"]
        c.border = style["border"]
    widths = [10, 50, 22, 28, 50, 30, 40, 70, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w
    ws.row_dimensions[1].height = 30


def append_row(wb: Workbook, sheet_name: str, row: dict) -> int:
    """Append one row; returns the 1-based Excel row number written."""
    ws = wb[sheet_name]
    next_row = ws.max_row + 1
    if ws.cell(row=1, column=1).value is None:
        next_row = 1
        _write_headers(ws)
        next_row = 2
    style = _row_style()
    values = [
        row.get("paper_id", ""),
        row.get("paper_title", ""),
        row.get("primary_name", ""),
        row.get("primary_email", ""),
        row.get("authors", ""),
        row.get("author_names", ""),
        row.get("author_emails", ""),
        row.get("review", ""),
        row.get("opinion", ""),
    ]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.alignment = style["alignment"]
        c.border = style["border"]
    return next_row


def save(wb: Workbook, path: str) -> None:
    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    wb.save(str(p))


def load_cmt_metadata(path: str) -> dict[str, dict]:
    """Read a CMT-export Excel and return {paper_id_str: {paper_title, primary_name, ...}}.

    Accepts:
    - real .xlsx workbooks (via openpyxl)
    - XML SpreadsheetML files saved with .xls extension (the format CMT actually exports)
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".xlsx":
        return _load_cmt_from_xlsx(path)
    if ext == ".xls":
        with open(path, "rb") as f:
            head = f.read(8)
        if head.startswith(b"<?xml"):
            return _load_cmt_from_xml(path)
        raise ValueError(
            "This looks like a binary .xls (older Excel). "
            "Open it in Excel and Save As .xlsx, then try again."
        )
    raise ValueError(f"Unsupported metadata file extension: {ext}")


def _load_cmt_from_xlsx(path: str) -> dict[str, dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, dict] = {}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_row = None
        col_to_field: dict[int, str] = {}
        for row in rows:
            if header_row is None:
                # find a header row that contains at least "Paper ID"
                norm = [str(c).strip() if c is not None else "" for c in row]
                if "Paper ID" in norm:
                    header_row = norm
                    for i, name in enumerate(norm):
                        if name in CMT_FIELDS:
                            col_to_field[i] = CMT_FIELDS[name]
                continue
            if not any(row):
                continue
            data: dict[str, str] = {}
            for i, val in enumerate(row):
                if i in col_to_field:
                    data[col_to_field[i]] = "" if val is None else str(val)
            pid = (data.get("paper_id") or "").strip()
            if pid:
                out[pid] = data
    wb.close()
    return out


def _load_cmt_from_xml(path: str) -> dict[str, dict]:
    NS = "urn:schemas-microsoft-com:office:spreadsheet"
    ns = {"ss": NS}
    with open(path, "r", encoding="utf-8") as f:
        root = ET.fromstring(f.read())
    out: dict[str, dict] = {}
    for ws in root.findall("ss:Worksheet", ns):
        table = ws.find("ss:Table", ns)
        if table is None:
            continue
        rows = table.findall("ss:Row", ns)
        header_cols: dict[int, str] = {}
        for ri, row in enumerate(rows):
            col = 0
            cells: dict[int, str] = {}
            for cell in row.findall("ss:Cell", ns):
                idx = cell.get(f"{{{NS}}}Index")
                if idx:
                    col = int(idx)
                else:
                    col += 1
                data = cell.find("ss:Data", ns)
                cells[col] = "".join(data.itertext()) if data is not None else ""
            if not header_cols:
                # try to identify header
                vals = list(cells.values())
                if any(v == "Paper ID" for v in vals):
                    for c, v in cells.items():
                        if v in CMT_FIELDS:
                            header_cols[c] = CMT_FIELDS[v]
                continue
            data_row: dict[str, str] = {}
            for c, v in cells.items():
                if c in header_cols:
                    data_row[header_cols[c]] = v
            pid = (data_row.get("paper_id") or "").strip()
            if pid:
                out[pid] = data_row
    return out


def delete_row(path: str, excel_row: int) -> bool:
    """Delete a single data row (1-based, where row 1 is the header).

    Returns True if a row was deleted. Raises if the file/row is invalid.
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() != ".xlsx":
        raise ValueError("Excel file not found or not an .xlsx")
    if excel_row < 2:
        raise ValueError("Row 1 is the header — cannot delete it.")
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    if excel_row > ws.max_row:
        wb.close()
        return False
    ws.delete_rows(excel_row, amount=1)
    wb.save(path)
    wb.close()
    return True


def delete_by_paper_id(path: str, paper_id: str) -> int:
    """Delete every data row whose Paper ID matches. Returns count removed."""
    p = Path(path)
    if not p.exists() or p.suffix.lower() != ".xlsx":
        raise ValueError("Excel file not found or not an .xlsx")
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    to_delete: list[int] = []
    for ri in range(2, ws.max_row + 1):
        v = ws.cell(row=ri, column=1).value
        if v is not None and str(v).strip() == str(paper_id).strip():
            to_delete.append(ri)
    # delete from bottom up so indices stay valid
    for ri in reversed(to_delete):
        ws.delete_rows(ri, amount=1)
    if to_delete:
        wb.save(path)
    wb.close()
    return len(to_delete)


def clear_all_rows(path: str) -> int:
    """Delete every data row but keep the header. Returns count removed."""
    p = Path(path)
    if not p.exists() or p.suffix.lower() != ".xlsx":
        raise ValueError("Excel file not found or not an .xlsx")
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    n = max(0, ws.max_row - 1)
    if n > 0:
        ws.delete_rows(2, amount=n)
        wb.save(path)
    wb.close()
    return n


def find_existing_ids(wb: Workbook, sheet_name: str) -> set[str]:
    """IDs already present in the workbook — used to avoid duplicate rows."""
    ws = wb[sheet_name]
    ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v is not None and str(v).strip():
            ids.add(str(v).strip())
    return ids
